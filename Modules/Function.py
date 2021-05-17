import openpyxl
from PyQt5.QtWidgets import QMessageBox

def CheckRegistrationDetails(user_id, user_name, password1, password2,full_name):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setWindowTitle("Error")

    #check user name
    if(not CheckUserName(user_name)):
        msg.setInformativeText('User Name Needs to be between 6 to 10 charcters and contains max 2 number')
        x = msg.exec_()
        return False

    #check password:
    if password1 != password2:
        print(password1 + password2)
        msg.setInformativeText("""Passwords Don't match""")
        msg.exec_()
        return False
    if(not CheckPassword(password1)):
        msg.setInformativeText('Password Length must be 8-10 contain at least 1 number  and 1 special charcter and 1 letter')
        msg.exec_()
        return False

    #check id
    if(not CheckID(user_id)):
        msg.setInformativeText('Invalid User id: length 9 charcter only numbers or Id already registered.')
        msg.exec_()
        return False

    #chack full name
    if(not CheckUserFullName(full_name)):
        msg.setInformativeText('Name Length at least 5 charcters long.')
        msg.exec_()
        return False

    Register(user_id, user_name, password1,full_name)
    return True


def CheckUserFullName(name):
    if len(name)<5:
        return False
    if not name.isalpha():
        return False
    return True


def CheckID(user_id):
    """
    :param user_id:
    :return: true if contains only digits and length=9 And the ID does not exist in the system
    """
    if not user_id.isdigit():
        return False
    if len(user_id) is not 9:
        return False
    ws = get_users_sheet().active
    for i in range (1,ws.max_row+1):
        id1=ws['A'+str(i)].value
        if id1==user_id:
            return False

    return True


def CheckUserName(user_name):
    """
    :param user_name:
    :return: True for correct username, otherwise false
    """
    # length between 6 to 8
    if len(user_name) > 8 or len(user_name) < 6:
        return False

    #Contains a maximum of 2 digits ,Contains letters or numbers
    count=0 #amount of digits
    for c in user_name:
        if c.isdigit():
            count+=1
        elif not c.isalpha(): #If the character is not a number it must be a letter
            return False
        if count>2:
            return False

    return True


def CheckPassword(password):
    if len(password)<8 or len(password)>10:
        return False

    count1=count2=count3=0
    for c in password:
        if c.isdigit():
            count1+=1
        elif c.isalpha():
            count2+=1
        else:
            count3+=1

        if(count1>=1 and count2>=1 and count3>=1):
            return True

    return False


def get_users_sheet():
    file = "users.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    return wb


def CheckLoginDetails(user_id, user_name, password):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setWindowTitle("Error")

    ws = get_users_sheet().active
    for i in range (1,ws.max_row+1):
        id1=ws['A'+str(i)].value
        name1=ws['B'+str(i)].value
        password1=ws['C'+str(i)].value

        if id1==user_id:
            if name1==user_name:
                if password==password1:
                    msg.setIcon(QMessageBox.Information)
                    msg.setText("Login Succefuly")
                    msg.setWindowTitle("Login Succefuly")
                    msg.exec_()
                    return True
                else:
                    msg.setInformativeText('Password is incorrect')
                    msg.exec_()
                    return False
            else:
                msg.setInformativeText('User Name is incorrect')
                msg.exec_()
                return False

    msg.setInformativeText('Id is incorrect')
    msg.exec_()
    return False


def Register(user_id,user_name,password1,full_name):
    ws = get_users_sheet()
    row=[user_id,user_name,password1,full_name]
    ws.active.append(row)
    ws.save("users.xlsx")


def CheckDictionaryValues(dict):
    """
    :param dict: Contains the patient's values
    :return: True if all values are non-negative numbers
    """
    names=list(dict.keys())
    values=list(dict.values())
    msg="Error - All values should be non-negative numbers. Incorrect values: "
    flag=True
    for i in range(0,len(values)):
        if not values[i].isnumeric(): #if its not a number / negative :"-"
            msg += names[i] + ","
            flag = False

    # if not flag: #ןf there are incorrect values- show mwssage
    # #show msg

    return flag


def Return_LOWorHIGHorNORMAL(a,b,val):
    if val < a:
        return "LOW"
    elif val > b:
        return "HIGH"
    else:
        return "NORMAL"


def WBC(val,age):
    if age >=18:
        a=4500
        b=11000
    elif age>=4:
       a=5500
       b=15500
    else:
        a=6000
        b=17500

    return Return_LOWorHIGHorNORMAL(a,b , val)


def Naut(val):
    return Return_LOWorHIGHorNORMAL(28,54,val)


def Lymph(val):
    return Return_LOWorHIGHorNORMAL(36,52,val)

def RBC(val):
    return Return_LOWorHIGHorNORMAL(4.5,6,val)

def HCT(val,gender):
    if gender== "F":
        a=33
        b=47
    else:
        a=37
        b=54
    return Return_LOWorHIGHorNORMAL(a,b,val)


def Urea(val,origin):
    if(origin):
        a=18.7
        b=47.3
    else:
        a=17
        b=43

    return Return_LOWorHIGHorNORMAL(a,b,val)


def Hb(val,age,gender):
    if age<=17:
        a=11.5
        b=15.5
    elif gender=='F':
        a=12
        b=16
    else:
        a=12
        b=18
    return Return_LOWorHIGHorNORMAL(a,b,val)


def Creatinine(val,age):
    if age<=2:
        a=0.2
        b=0.5
    elif age<=17:
        a=0.5
        b=1
    elif age<=59:
        a=0.6
        b=1
    else:
        a=0.6
        b=1.2
    return Return_LOWorHIGHorNORMAL(a,b,val)


def Iron(val,gender):
    if gender=='F':
        a=48
        b=128
    else:
        a=60
        b=160
    return Return_LOWorHIGHorNORMAL(a,b,val)


def HDL(val,gender,origin):
    if gender=='F':
        a=34
        b=82
    else:
        a=29
        b=61
    if(origin):
        a*=1.2
        b*=1.2
    return Return_LOWorHIGHorNORMAL(a,b,val)


def AP(val,origin):
    if origin:
        a=60
        b=120
    else:
        a=30
        b=90
    return Return_LOWorHIGHorNORMAL(a,b,val)


def ConvertsValuesTo_LOW_HIGH_NORMAL(dict):
    age=dict["age"]
    gender=dict["gender"]
    origin=dict["origin"]

    dict["WBC"]=WBC(dict["WBC"],age)
    dict["Neut"]=Naut(dict["Neut"])
    dict["Lymph"]=Lymph(dict["Lymph"])
    dict["RBC"]=RBC(dict["RBC"])
    dict["HCT"]=HCT(dict["HCT"],gender)
    dict["Urea"]=Urea(dict["Urea"],origin)
    dict["Hb"]=Hb(dict["Hb"],age,gender)
    dict["Creatinine"]=Creatinine(dict["Creatinine"],age)
    dict["Iron"]=Iron(dict["Iron"],gender)
    dict["HDL"]=HDL(dict["HDL"],gender,origin)
    dict["AP"]=AP(dict["AP"],origin)


def Treatment_according_to_diagnosis(diagnosis):
    dict={"Anemia":"Two pills of 10 mg of B12 a day for a month",
          "Diet":"To coordinate an appointment with a nutritionist",
          "Bleeding":"To be evacuated urgently to the hospital",
          "Hyperlipidemia (blood lipids)":"schedule an appointment with a nutritionist, 5 mg of Simobil pill per day for a week",
          "Disorder of blood formation / blood cells":"10 mg pill of B12 and 5 mg pill of folic acid per day for one month",
          "Hematologic disorder":"injection of a hormone to encourage red blood cell production",
          "Iron poisoning": "to evacuate to hospital",
          "Dehydration":"complete rest when lying down, returning fluids to drinking",
          "Infection":"dedicated antibiotics",
          "Vitamin deficiency":"referral for a blood test to identify the missing vitamins",
          "Viral disease":"rest at home",
          "Bile duct diseases":"referral to surgical treatment",
          "Heart disease":"make an appointment with a nutritionist",
          "Blood disease":"a combination of cyclophosphamide and corticosteroids",
          "Liver disease":"referral to a specific diagnosis for treatment",
          "Kidney disease":"balancing blood sugar levels",
          "Iron deficiency":"two 10 mg pills of B12 a day for a month",
          "Muscle diseases":"two 5 mg pills of Altman c3 turmeric a day for a month",
          "Smokers":"quit smoking",
          "Lung disease":"stop smoking / refer to X-ray of the lungs",
          "Hypothyroidism":"Propylthiouracil to reduce thyroid activity",
          "Adult diabetes":"insulin adjustment for the patient",
          "Cancer":"Entrectinib",
          "Increased consumption of meat":"To coordinate an appointment with a nutritionist",
          "Use of various drugs":"Referral to a family doctor for a match between medications",
          "Malnutrition":"To coordinate an appointment with a nutritionist"

    }

    return dict[diagnosis]


def get_diagnosis_dict(person):
    diagnosis={"Anemia":0,"Diet":0,"Bleeding":0,"Hyperlipidemia (blood lipids)":0,"Disorder of blood formation / blood cells":0,"Hematologic disorder":0,"Iron poisoning":0,
               "Dehydration":0,"Infection":0,"Vitamin deficiency":0,"Viral disease":0,"Bile duct diseases":0,"Heart disease":0,"Blood disease":0,"Liver disease":0,"Kidney disease":0,
               "Iron deficiency":0,"Muscle diseases":0,"Smokers":0,"Lung disease":0,"Hypothyroidism":0,"Adult diabetes":0,"Cancer":0,"Increased consumption of meat":0,
               "Use of various drugs":0,"Malnutrition":0}

    WBCdiagnosis(person, diagnosis)
    Nautdiagnosis(person, diagnosis)
    Lymphdiagnosis(person, diagnosis)
    RBCdiagnosis(person, diagnosis)
    HCTdiagnosis(person, diagnosis)
    Ureadiagnosis(person, diagnosis)
    Hbdiagnosis(person, diagnosis)
    Creatininediagnosis(person, diagnosis)
    Irondiagnosis(person, diagnosis)
    HDLdiagnosis(person, diagnosis)
    APdiagnosis(person, diagnosis)

    return diagnosis


def get_string_of_diagnosis_and_Treatment(diagnosis):
    max = max(diagnosis.values())
    if max==0:
        return "The tests are normal and you are a healthy person."
    Main_diagnoses = list(map(lambda x: x[0], (filter(lambda x: x[1] == max, diagnosis.items()))))
    Secondary_diagnoses=list(map(lambda x: x[0], (filter(lambda x: x[1] == max-1, diagnosis.items()))))
    ans=""
    for i in Main_diagnoses:
        ans+="Diagnoses: "+i+"\n" + "Treatment: "+Treatment_according_to_diagnosis(i)+"\n"

    if(len(Secondary_diagnoses)>0):
        ans+="\nIn addition there are concerns:\n"
        for i in Secondary_diagnoses:
            ans += "Diagnoses: " + i + "\n" + "Treatment: " + Treatment_according_to_diagnosis(i) + "\n"


def WBCdiagnosis(person,diagnosis):
    if(person["WBC"]=="HIGH"):
        if person["Fever"]:
            diagnosis["Infection"]+=1
        else:
            diagnosis["Blood disease"] += 0.5
            diagnosis["Cancer"]+=0.5
    elif(person["WBC"]=="LOW"):
        diagnosis["Cancer"] += 0.5
        diagnosis["Viral disease"] += 1


def Nautdiagnosis(person,diagnosis):
    if (person["Naut"] == "HIGH"):
        diagnosis["Infection"] += 1
    elif (person["Naut"] == "LOW"):
        diagnosis["Cancer"] += 0.5
        diagnosis["Infection"] += 1


def Lymphdiagnosis(person,diagnosis):
    if (person["Naut"] == "HIGH"):
        diagnosis["Infection"] += 1
        diagnosis["Cancer"] += 1
    elif (person["Naut"] == "LOW"):
        diagnosis["Disorder of blood formation / blood cells"] += 1


def RBCdiagnosis(person,diagnosis):
    if person["RBC"] == "HIGH":
        diagnosis["Disorder of blood formation / blood cells"] += 1
        diagnosis["Lung disease"] += 1
        if person["smoker"]:
            diagnosis["Smokers"] += 1
    elif (person["RBC"] == "LOW"):
        diagnosis["Anemia"] += 1
        diagnosis["Bleeding"] += 1


def HCTdiagnosis(person,diagnosis):
    if person["HCT"] == "HIGH":
        if person["smoker"]:
            diagnosis["Smokers"] += 1
    elif (person["HCT"] == "LOW"):
        diagnosis["Anemia"] += 1
        diagnosis["Bleeding"] += 1


def Ureadiagnosis(person,diagnosis):
    if person["Urea"] == "HIGH":
        diagnosis["Kidney disease"] += 1
        diagnosis["Dehydration"] += 1
        diagnosis["Diet"] += 1
    elif (person["Urea"] == "LOW") and not person["pregnancy"]:
        diagnosis["Malnutrition"] += 1
        diagnosis["Diet"] += 1
        diagnosis["Liver disease"] += 1


def Hbdiagnosis(person,diagnosis):
    if person["Hb"] == "LOW" :
        diagnosis["Anemia"] += 1
        diagnosis["Hematologic disorder"] += 1
        diagnosis["Iron deficiency"] += 1
        diagnosis["Bleeding"] += 1


def Creatininediagnosis(person,diagnosis):
    if person["Creatinine"] == "HIGH":
        diagnosis["Kidney disease"] += 1
        diagnosis["Muscle diseases"] += 1
        diagnosis["Increased consumption of meat"] += 1
    elif (person["Creatinine"] == "LOW") :
        diagnosis["Malnutrition"] += 1


def Irondiagnosis(person,diagnosis):
    if person["Iron"] == "HIGH":
        diagnosis["Iron poisoning"] += 1
    elif (person["Iron"] == "LOW"):
        diagnosis["Iron deficiency"] += 1
        if not person["pregnancy"]:
            diagnosis["Bleeding"] += 1
            diagnosis["Malnutrition"] += 1


def HDLdiagnosis(person,diagnosis):
    if (person["HDL"] == "LOW"):
        diagnosis["Heart disease"] += 1
        diagnosis["Hyperlipidemia (blood lipids)"] += 1
        diagnosis["Adult diabetes"] += 1


def APdiagnosis(person,diagnosis):
    if person["AP"] == "HIGH" and not person["pregnancy"]:
        diagnosis["Liver disease"] += 1
        diagnosis["Bile duct diseases"] += 1
        diagnosis["Hypothyroidism"] += 1
        diagnosis["Use of various drugs"] += 1
    elif (person["AP"] == "LOW") :
        diagnosis["Vitamin deficiency"] += 1
        diagnosis["Malnutrition"] += 1


